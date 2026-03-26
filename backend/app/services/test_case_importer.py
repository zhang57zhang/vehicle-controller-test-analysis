"""
测试用例导入服务

提供测试用例Excel导入和导出功能
"""

import logging
import io
from typing import Dict, Any, List, Optional
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows

from app.config import settings

logger = logging.getLogger(__name__)


class TestCaseImportError(Exception):
    """测试用例导入错误"""

    pass


class TestCaseImporter:
    """
    测试用例导入器

    支持Excel格式的测试用例导入和导出
    """

    REQUIRED_COLUMNS = ["TC_ID", "TC_Name", "Test_Phase"]
    OPTIONAL_COLUMNS = [
        "Pre_Condition",
        "Test_Steps",
        "Expected_Result",
        "Priority",
        "Version",
        "Author",
        "Notes",
    ]

    COLUMN_MAPPING = {
        "TC_ID": "tc_id",
        "TC_Name": "tc_name",
        "Test_Phase": "test_phase",
        "Pre_Condition": "pre_condition",
        "Test_Steps": "test_steps",
        "Expected_Result": "expected_result",
        "Priority": "priority",
        "Version": "version",
        "Author": "author",
        "Notes": "notes",
        "tc_id": "tc_id",
        "tc_name": "tc_name",
        "test_phase": "test_phase",
        "pre_condition": "pre_condition",
        "test_steps": "test_steps",
        "expected_result": "expected_result",
        "priority": "priority",
        "version": "version",
        "author": "author",
        "notes": "notes",
    }

    VALID_TEST_PHASES = ["MIL", "HIL", "DVP", "VEHICLE", "MANUAL", "AUTO"]
    VALID_PRIORITIES = ["High", "Medium", "Low", "high", "medium", "low"]

    def parse_excel(self, file_path: str) -> List[Dict[str, Any]]:
        """
        解析Excel文件中的测试用例

        Args:
            file_path: Excel文件路径

        Returns:
            测试用例列表
        """
        try:
            df = pd.read_excel(file_path, engine="openpyxl")

            df.columns = [col.strip() for col in df.columns]

            missing_columns = [col for col in self.REQUIRED_COLUMNS if col not in df.columns]
            if missing_columns:
                alt_mapping = {
                    "TC_ID": ["tc_id", "TestCaseID", "Test Case ID", "用例编号", "编号"],
                    "TC_Name": ["tc_name", "TestCaseName", "Test Case Name", "用例名称", "名称"],
                    "Test_Phase": ["test_phase", "TestPhase", "Test Phase", "测试阶段", "阶段"],
                }

                for req_col in missing_columns:
                    for alt_col in alt_mapping.get(req_col, []):
                        if alt_col in df.columns:
                            df.rename(columns={alt_col: req_col}, inplace=True)
                            break

                missing_columns = [col for col in self.REQUIRED_COLUMNS if col not in df.columns]
                if missing_columns:
                    raise TestCaseImportError(
                        f"Missing required columns: {', '.join(missing_columns)}. "
                        f"Required columns: {self.REQUIRED_COLUMNS}"
                    )

            test_cases = []
            errors = []

            for idx, row in df.iterrows():
                try:
                    tc = self._parse_row(row, idx + 2)
                    if tc:
                        test_cases.append(tc)
                except Exception as e:
                    errors.append(f"Row {idx + 2}: {str(e)}")

            if errors:
                logger.warning(f"Import completed with {len(errors)} errors: {errors[:5]}")

            logger.info(f"Parsed {len(test_cases)} test cases from {file_path}")

            return {
                "test_cases": test_cases,
                "total_rows": len(df),
                "imported_count": len(test_cases),
                "error_count": len(errors),
                "errors": errors[:10],
            }

        except TestCaseImportError:
            raise
        except Exception as e:
            logger.error(f"Error parsing Excel file: {str(e)}", exc_info=True)
            raise TestCaseImportError(f"Failed to parse Excel file: {str(e)}")

    def _parse_row(self, row: pd.Series, row_num: int) -> Optional[Dict[str, Any]]:
        """解析单行数据"""
        tc_id = str(row.get("TC_ID", "")).strip()
        tc_name = str(row.get("TC_Name", "")).strip()
        test_phase = str(row.get("Test_Phase", "")).strip().upper()

        if not tc_id or tc_id == "nan":
            return None

        if not tc_name or tc_name == "nan":
            raise ValueError(f"TC_Name is required for TC_ID: {tc_id}")

        if test_phase not in self.VALID_TEST_PHASES:
            test_phase = "HIL"

        priority = str(row.get("Priority", "Medium")).strip()
        if priority.lower() not in ["high", "medium", "low"]:
            priority = "Medium"
        priority = priority.capitalize()

        tc = {
            "tc_id": tc_id,
            "tc_name": tc_name,
            "test_phase": test_phase,
            "pre_condition": self._safe_str(row.get("Pre_Condition")),
            "test_steps": self._safe_str(row.get("Test_Steps")),
            "expected_result": self._safe_str(row.get("Expected_Result")),
            "priority": priority,
            "version": self._safe_str(row.get("Version", "1.0")),
            "author": self._safe_str(row.get("Author")),
            "notes": self._safe_str(row.get("Notes")),
            "status": "active",
        }

        return tc

    def _safe_str(self, value) -> str:
        """安全转换为字符串"""
        if pd.isna(value) or value is None:
            return ""
        return str(value).strip()

    def generate_template(self, output_path: str) -> str:
        """
        生成测试用例导入模板

        Args:
            output_path: 输出文件路径

        Returns:
            生成的模板文件路径
        """
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "测试用例"

            headers = [
                "TC_ID",
                "TC_Name",
                "Test_Phase",
                "Pre_Condition",
                "Test_Steps",
                "Expected_Result",
                "Priority",
                "Version",
                "Author",
                "Notes",
            ]

            header_fill = PatternFill(start_color="1890FF", end_color="1890FF", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            header_alignment = Alignment(horizontal="center", vertical="center")
            thin_border = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin"),
            )

            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_alignment
                cell.border = thin_border

            sample_data = [
                [
                    "TC_001",
                    "电机启动测试",
                    "HIL",
                    "系统已上电",
                    "1. 发送启动指令\n2. 检查电机状态",
                    "电机转速 > 0",
                    "High",
                    "1.0",
                    "张三",
                    "",
                ],
                [
                    "TC_002",
                    "电压边界测试",
                    "HIL",
                    "系统正常运行",
                    "1. 降低输入电压\n2. 检查欠压保护",
                    "触发欠压故障",
                    "High",
                    "1.0",
                    "李四",
                    "",
                ],
                [
                    "TC_003",
                    "温度保护测试",
                    "HIL",
                    "系统正常运行",
                    "1. 升高温度\n2. 检查过温保护",
                    "触发过温保护",
                    "Medium",
                    "1.0",
                    "王五",
                    "",
                ],
            ]

            for row_idx, row_data in enumerate(sample_data, 2):
                for col_idx, value in enumerate(row_data, 1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    cell.border = thin_border
                    cell.alignment = Alignment(vertical="top", wrap_text=True)

            ws.column_dimensions["A"].width = 12
            ws.column_dimensions["B"].width = 25
            ws.column_dimensions["C"].width = 12
            ws.column_dimensions["D"].width = 20
            ws.column_dimensions["E"].width = 30
            ws.column_dimensions["F"].width = 25
            ws.column_dimensions["G"].width = 10
            ws.column_dimensions["H"].width = 10
            ws.column_dimensions["I"].width = 12
            ws.column_dimensions["J"].width = 20

            ws.row_dimensions[1].height = 25
            for row in range(2, len(sample_data) + 2):
                ws.row_dimensions[row].height = 60

            wb.save(output_path)

            logger.info(f"Generated test case template: {output_path}")
            return output_path

        except Exception as e:
            logger.error(f"Error generating template: {str(e)}", exc_info=True)
            raise TestCaseImportError(f"Failed to generate template: {str(e)}")

    def export_test_results(
        self, test_results: List[Dict[str, Any]], output_path: str, format: str = "excel"
    ) -> str:
        """
        导出测试结果

        Args:
            test_results: 测试结果列表
            output_path: 输出文件路径
            format: 输出格式 ('excel' 或 'csv')

        Returns:
            导出的文件路径
        """
        try:
            if format == "csv":
                df = pd.DataFrame(test_results)
                df.to_csv(output_path, index=False, encoding="utf-8-sig")
            else:
                wb = Workbook()
                ws = wb.active
                ws.title = "测试结果"

                headers = [
                    "TC_ID",
                    "TC_Name",
                    "Test_Phase",
                    "Result",
                    "Signal_Name",
                    "Measured_Value",
                    "Expected_Min",
                    "Expected_Max",
                    "Pass/Fail",
                    "Notes",
                    "Executed_At",
                ]

                header_fill = PatternFill(
                    start_color="52C41A", end_color="52C41A", fill_type="solid"
                )
                header_font = Font(bold=True, color="FFFFFF")
                thin_border = Border(
                    left=Side(style="thin"),
                    right=Side(style="thin"),
                    top=Side(style="thin"),
                    bottom=Side(style="thin"),
                )

                for col, header in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=col, value=header)
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    cell.border = thin_border

                pass_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                fail_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                warn_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")

                for row_idx, result in enumerate(test_results, 2):
                    values = [
                        result.get("tc_id", ""),
                        result.get("tc_name", ""),
                        result.get("test_phase", ""),
                        result.get("result", ""),
                        result.get("signal_name", ""),
                        result.get("measured_value", ""),
                        result.get("expected_min", ""),
                        result.get("expected_max", ""),
                        result.get("pass_fail", ""),
                        result.get("notes", ""),
                        result.get("executed_at", ""),
                    ]

                    for col_idx, value in enumerate(values, 1):
                        cell = ws.cell(
                            row=row_idx, column=col_idx, value=str(value) if value else ""
                        )
                        cell.border = thin_border

                        if col_idx == 9:
                            if str(value).upper() == "PASS":
                                cell.fill = pass_fill
                            elif str(value).upper() == "FAIL":
                                cell.fill = fail_fill
                            else:
                                cell.fill = warn_fill

                for col_letter, width in [
                    ("A", 12),
                    ("B", 25),
                    ("C", 12),
                    ("D", 10),
                    ("E", 20),
                    ("F", 15),
                    ("G", 12),
                    ("H", 12),
                    ("I", 10),
                    ("J", 25),
                    ("K", 20),
                ]:
                    ws.column_dimensions[col_letter].width = width

                wb.save(output_path)

            logger.info(f"Exported {len(test_results)} test results to {output_path}")
            return output_path

        except Exception as e:
            logger.error(f"Error exporting test results: {str(e)}", exc_info=True)
            raise TestCaseImportError(f"Failed to export test results: {str(e)}")


test_case_importer = TestCaseImporter()
